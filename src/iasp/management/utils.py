import logging
import os
import logging

from django.conf import settings
from django.http import HttpResponse

from applications.models import *
from calls.models import CallFreeCreditsRule

from openpyxl import Workbook
from openpyxl.styles import Font


logger = logging.getLogger(__name__)


def export_xls(application):
    # Crea un nuovo file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Dati"

    ws.append(["Candidato", str(application.user)])
    ws.append(["Nazionalità", application.user_country])
    ws.append(["Università", f"{application.home_university} ({application.home_city} - {application.home_country})"])
    ws.append(["Corso", application.home_course])

    ws.append([])

    ws.append(["Data di invio", application.submission_date.strftime("%d/%m/%Y %H:%M:%S")])
    ws.append(["Num. protocollo", application.protocol_number])
    ws.append(["Data protocollo", application.protocol_date.strftime("%d/%m/%Y %H:%M:%S") if application.protocol_date else "-"])

    # required
    insertions_required = ApplicationInsertionRequired.objects.filter(
        application=application
    ).prefetch_related('review').order_by('target_teaching_name')

    if insertions_required.exists():
        ws.append([])

        ws.append(["Insegnamenti previsti dal piano"])
        # Applica il grassetto
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

        ws.append([])

        labels = [
            "Attività formativa sostenuta"
        ]

        if not application.call.insertions_only_from_same_course:
            labels.extend(
                [
                    "Università",
                    "Corso",
                ]
            )

        labels.extend(
            [
                "CFU",
                "Voto",
                "Attività formativa convalidata",
                "CFU riconosciuti",
                "Voto riconosciuto",
                "Note",
            ]
        )

        ws.append(labels)

        # Applica il grassetto all'ultima riga
        for col in range(1, len(labels) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

        for required in insertions_required:
            data = [
                f"{required.source_teaching_name} ({required.source_teaching_ssd or '-'})",
            ]

            if not application.call.insertions_only_from_same_course:
                data.extend(
                    [
                        f"{required.source_university} ({required.source_university_city} - {required.source_university_country})",
                        required.source_degree_course
                    ]
                )

            data.extend(
                [
                    required.source_teaching_credits,
                    required.source_teaching_grade,
                    f"{required.target_teaching_cod} - {required.target_teaching_name} - {required.target_teaching_ssd} ({required.target_teaching_credits} CFU)".replace("\r", "").replace("\n", ""),
                    required.review.changed_credits if hasattr(required, 'review') else required.source_teaching_credits,
                    required.review.changed_grade if hasattr(required, 'review') else required.source_teaching_grade,
                    required.review.notes if hasattr(required, 'review') else "-",
                ]
            )

            ws.append(data)
    # end required

    # free
    insertions_free = ApplicationInsertionFree.objects.filter(
        application=application
    ).select_related('free_credits')

    if insertions_free.exists():
        ws.append([])
        ws.append(["Insegnamenti a scelta"])
        # Applica il grassetto
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append([])

        labels = [
            "Attività formativa sostenuta",
        ]

        if not application.call.insertions_only_from_same_course:
            labels.extend(
                [
                    "Università",
                    "Corso",
                ]
            )

        labels.extend(
            [
                "CFU",
                "Voto",
                "Vincolo insegnamenti a scelta",
                "CFU riconosciuti",
                "Voto riconosciuto",
                "Note",
            ]
        )

        ws.append(labels)

        # Applica il grassetto all'ultima riga
        for col in range(1, len(labels) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

        for free in insertions_free:
            data = [
                f"{free.source_teaching_name} ({free.source_teaching_ssd or '-'})",
            ]

            if not application.call.insertions_only_from_same_course:
                data.extend(
                    [
                        f"{free.source_university} ({free.source_university_city} - {free.source_university_country})",
                        required.source_degree_course
                    ]
                )

            data.extend(
                [
                    free.source_teaching_credits,
                    free.source_teaching_grade,
                    f"Insegnamenti a scelta {free.free_credits.course_year}° anno (max {free.free_credits.max_value} CFU)".replace("\r", "").replace("\n", ""),
                    free.review.changed_credits if hasattr(free, 'review') else free.source_teaching_credits,
                    free.review.changed_grade if hasattr(free, 'review') else free.source_teaching_grade,
                    free.review.notes if hasattr(free, 'review') else "-",
                ]
            )

            ws.append(data)
    # end free

    # Prepara la risposta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="domanda_{application.pk}.xlsx"'

    # Salva il file direttamente nella response
    wb.save(response)

    return response
