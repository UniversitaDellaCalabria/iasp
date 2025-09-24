import logging
import magic
import os
import pypdf
import shutil

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import PermissionDenied
from django.core.mail import send_mail
from django.db.models import Q
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_POST

from applications.decorators import activity_check
from applications.forms import InsertionRequiredForm, PaymentForm
from applications.models import *
from applications.utils import *

from calls.models import *

from django_tables2 import RequestConfig

from generics.models import Log

from openpyxl import Workbook
from openpyxl.styles import Font

from organizational_area.decorators import *

from pathlib import Path

from .. decorators import *
from .. filters import ApplicationFilter
from .. forms import *
from .. models import *
from .. settings import *
from .. tables import *


logger = logging.getLogger(__name__)



@login_required
@belongs_to_a_commission
def list(request, commissions=None):
    template = 'commissions/list.html'

    return render(
        request,
        template,
        {
            'commissions': commissions,
        }
    )


@login_required
@belongs_to_commission
def detail(request, call_pk, commission=None):
    template = 'commissions/detail.html'
    return render(
        request,
        template,
        {
            'call': commission.call,
            'commission': commission,
        }
    )


@login_required
@belongs_to_commission
def applications(request, call_pk, commission=None):
    template = 'commissions/applications.html'
    applications = Application.objects.filter(
        call=commission.call,
        submission_date__isnull=False
    )
    f = ApplicationFilter(request.GET, queryset=applications)
    table = CommissionApplicationTable(f.qs)
    RequestConfig(request, paginate={"per_page": APPLICATIONS_PAGINATION}).configure(table)

    return render(
        request,
        template,
        {
            'commission': commission,
            'filter': f,
            'table': table
        }
    )


@login_required
@belongs_to_commission
@application_check
def application(request, call_pk, application_pk, commission=None, application=None):
    template = 'commissions/application.html'

    tot_credits = application.get_credits_status(show_commission_review=True)

    insertions_required = ApplicationInsertionRequired.objects.filter(
        application=application
    ).count()

    free_credits_rules = CallFreeCreditsRule.objects.filter(
        call__pk=call_pk,
        is_active=True
    )
    insertions_free = ApplicationInsertionFree.objects.filter(
        application=application
    ).count()

    form = PaymentForm(instance=application)

    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            'form': form,
            'free_credits_rules': free_credits_rules,
            'insertions_required': insertions_required,
            'insertions_free': insertions_free,
            'tot_credits': tot_credits
        }
    )


@login_required
@belongs_to_commission
@application_check
def application_required_list(request, call_pk, application_pk, commission=None, application=None):
    template = 'commissions/application_required_list.html'

    application_data = get_application_required_insertions_data(
        application=application,
        show_commission_review=True
    )

    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            **application_data
        }
    )


@login_required
@belongs_to_commission
@application_check
@activity_check
def application_required(request, call_pk, application_pk, teaching_id, commission=None, application=None, target_teaching={}):
    insertions = ApplicationInsertionRequired.objects.filter(
        application_id=application_pk,
        target_teaching_id=teaching_id
    )

    template = 'commissions/application_required.html'
    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            'insertions': insertions,
            'target_teaching': target_teaching
        }
    )


@login_required
@belongs_to_commission
@application_check
@activity_check
def application_required_detail(request, call_pk, application_pk, teaching_id, insertion_pk, commission=None, application=None, target_teaching={}):
    template = 'commissions/application_required_detail.html'
    insertion = get_object_or_404(
        ApplicationInsertionRequired,
        application=application,
        pk=insertion_pk
    )

    form = InsertionRequiredForm(
        target_teaching=target_teaching,
        instance=insertion,
        application=application
    )

    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            'form': form,
            'target_teaching': target_teaching
        }
    )


@login_required
@belongs_to_commission
@application_check
def application_free(request, call_pk, application_pk, year, commission=None, application=None):
    data = get_application_free_insertions_data(
        application=application,
        year=year,
        show_commission_review=True
    )
    template = 'commissions/application_free.html'
    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            **data
        }
    )


@login_required
@belongs_to_commission
@application_check
@application_check
def application_free_detail(request, call_pk, application_pk, year, insertion_pk, commission=None, application=None):
    free_credits_rule = get_object_or_404(
        CallFreeCreditsRule,
        is_active=True,
        call=application.call,
        course_year=year
    )

    insertion = get_object_or_404(
        ApplicationInsertionFree,
        application=application,
        pk=insertion_pk
    )

    form = InsertionFreeForm(
        instance=insertion,
        application=application,
        free_credits_rule=free_credits_rule
    )
    template = 'commissions/application_free_detail.html'

    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            'free_credits_rule': free_credits_rule,
            'form': form,
        }
    )


@login_required
@belongs_to_commission
@application_check
@activity_check
def application_required_review(request, call_pk, application_pk, teaching_id, insertion_pk, commission=None, application=None, target_teaching={}):
    template = 'commissions/application_required_review.html'
    insertion = get_object_or_404(
        ApplicationInsertionRequired,
        application=application,
        pk=insertion_pk
    )

    review = getattr(insertion, "review", None)

    form = ApplicationInsertionRequiredCommissionReviewForm(
        instance=review,
    )

    if request.method == 'POST':

        form = ApplicationInsertionRequiredCommissionReviewForm(
            instance=review,
            data=request.POST,
        )

        if form.is_valid():
            review = form.save(commit=False)
            review.insertion = insertion
            review.modified_by = request.user
            review.save()

            Log.objects.create(
                created_by=request.user,
                content_type = ContentType.objects.get_for_model(insertion),
                object_id=insertion.pk,
                text="{}: {} / {}: {} / {}: {}".format(
                    _("New credits value"),
                    review.changed_credits,
                    _("New grade value"),
                    review.changed_grade,
                    _("Notes"),
                    review.notes
                )
            )

            # log
            logger.info(
                "[{time}] utente {user}, commissione \"{commission}\" del bando \"{call}\", "
                "ha revisionato il valore dei CFU (da {old_credits} a {new_credits}), "
                "voto da {old_grade} a {new_grade}, "
                "per l'insegnamento di \"{teaching}\", inserito per la convalida di {target_teaching}, "
                "nella domanda di {student}".format(
                    time=timezone.localtime(),
                    user=request.user,
                    commission=commission.name,
                    call=commission.call,
                    old_credits=insertion.source_teaching_credits,
                    old_grade=insertion.source_teaching_grade,
                    new_credits=review.changed_credits,
                    new_grade=review.changed_grade,
                    teaching=insertion.source_teaching_name,
                    target_teaching=insertion.target_teaching_name,
                    student=application.user,
                )
            )
            # end log

            messages.add_message(
                request,
                messages.SUCCESS,
                _('Data modified successfully')
            )
            return redirect(
                'management:commission_application_required',
                call_pk=call_pk,
                application_pk=application_pk,
                teaching_id=teaching_id
            )
        else:  # pragma: no cover
            messages.add_message(
                request,
                messages.ERROR,
                "<b>{}</b>: {}".format(_('Warning'), _('There are errors in the form'))
            )
    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            'form': form,
            'insertion': insertion,
            'target_teaching': target_teaching,
        }
    )


@login_required
@belongs_to_commission
@application_check
def application_free_review(request, call_pk, application_pk, year, insertion_pk, commission=None, application=None, target_teaching={}):
    template = 'commissions/application_free_review.html'

    free_credits_rule = get_object_or_404(
        CallFreeCreditsRule,
        is_active=True,
        call=application.call,
        course_year=year
    )

    insertion = get_object_or_404(
        ApplicationInsertionFree,
        application=application,
        pk=insertion_pk
    )

    review = getattr(insertion, "review", None)

    form = ApplicationInsertionFreeCommissionReviewForm(
        instance=review,
    )

    if request.method == 'POST':

        form = ApplicationInsertionFreeCommissionReviewForm(
            instance=review,
            data=request.POST,
        )

        if form.is_valid():
            review = form.save(commit=False)
            review.insertion = insertion
            review.modified_by = request.user
            review.save()

            Log.objects.create(
                created_by=request.user,
                content_type = ContentType.objects.get_for_model(insertion),
                object_id=insertion.pk,
                text="{}: {} / {}: {} / {}: {}".format(
                    _("New credits value"),
                    review.changed_credits,
                    _("New grade value"),
                    review.changed_grade,
                    _("Notes"),
                    review.notes
                )
            )

            # log
            logger.info(
                "[{time}] utente {user}, commissione \"{commission}\" del bando \"{call}\", "
                "ha revisionato il valore dei CFU (da {old_credits} a {new_credits}), "
                "voto da {old_grade} a {new_grade}, "
                "per l'insegnamento di \"{teaching}\", crediti a scelta {year}° anno, "
                "nella domanda di {student}".format(
                    time=timezone.localtime(),
                    user=request.user,
                    commission=commission.name,
                    call=commission.call,
                    old_credits=insertion.source_teaching_credits,
                    old_grade=insertion.source_teaching_grade,
                    new_credits=review.changed_credits,
                    new_grade=review.changed_grade,
                    teaching=insertion.source_teaching_name,
                    year=insertion.free_credits.course_year,
                    student=application.user,
                )
            )
            # end log

            messages.add_message(
                request,
                messages.SUCCESS,
                _('Data modified successfully')
            )
            return redirect(
                'management:commission_application_free',
                call_pk=call_pk,
                application_pk=application_pk,
                year=year
            )
        else:  # pragma: no cover
            messages.add_message(
                request,
                messages.ERROR,
                "<b>{}</b>: {}".format(_('Warning'), _('There are errors in the form'))
            )
    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            'form': form,
            'insertion': insertion,
            'year': year
        }
    )


@login_required
@require_POST
@belongs_to_commission
@application_check
@activity_check
def application_required_review_delete(request, call_pk, application_pk, teaching_id, insertion_pk, commission=None, application=None, target_teaching={}):
    insertion = get_object_or_404(
        ApplicationInsertionRequired,
        application=application,
        pk=insertion_pk
    )

    review = getattr(insertion, "review", None)

    if not review:
        messages.add_message(
            request,
            messages.ERROR,
            _('Review not found')
        )
    else:
        Log.objects.create(
            created_by=request.user,
            content_type = ContentType.objects.get_for_model(insertion),
            object_id=insertion.pk,
            text=_("Review deleted")
        )

        # log
        logger.info(
            "[{time}] utente {user}, commissione \"{commission}\" del bando \"{call}\", "
            "ha eliminato la revisione dei CFU "
            "per l'insegnamento di \"{teaching}\", inserito per la convalida di {target_teaching}, "
            "nella domanda di {student}".format(
                time=timezone.localtime(),
                user=request.user,
                commission=commission.name,
                call=commission.call,
                teaching=insertion.source_teaching_name,
                target_teaching=insertion.target_teaching_name,
                student=application.user,
            )
        )
        # end log

        review.delete()

        messages.add_message(
            request,
            messages.SUCCESS,
            _('Review removed successfully')
        )

    return redirect(
        'management:commission_application_required',
        call_pk=call_pk,
        application_pk=application_pk,
        teaching_id=teaching_id
    )


@login_required
@require_POST
@belongs_to_commission
@application_check
def application_free_review_delete(request, call_pk, application_pk, year, insertion_pk, commission=None, application=None, target_teaching={}):
    free_credits_rule = get_object_or_404(
        CallFreeCreditsRule,
        is_active=True,
        call=application.call,
        course_year=year
    )

    insertion = get_object_or_404(
        ApplicationInsertionFree,
        application=application,
        pk=insertion_pk
    )

    review = getattr(insertion, "review", None)

    if not review:
        messages.add_message(
            request,
            messages.ERROR,
            _('Review not found')
        )
    else:
        Log.objects.create(
            created_by=request.user,
            content_type = ContentType.objects.get_for_model(insertion),
            object_id=insertion.pk,
            text=_("Review deleted")
        )

        # log
        logger.info(
            "[{time}] utente {user}, commissione \"{commission}\" del bando \"{call}\", "
            "ha eliminato la revisione dei CFU "
            "per l'insegnamento di \"{teaching}\", crediti a scelta {year}° anno, "
            "nella domanda di {student}".format(
                time=timezone.localtime(),
                user=request.user,
                commission=commission.name,
                call=commission.call,
                teaching=insertion.source_teaching_name,
                year=insertion.free_credits.course_year,
                student=application.user,
            )
        )
        # end log

        review.delete()

        messages.add_message(
            request,
            messages.SUCCESS,
            _('Review removed successfully')
        )

    return redirect(
        'management:commission_application_free',
        call_pk=call_pk,
        application_pk=application_pk,
        year=year
    )


@login_required
@belongs_to_commission
@application_check
@activity_check
def application_required_review_logs(request, call_pk, application_pk, teaching_id, insertion_pk, commission=None, application=None, target_teaching={}):
    template = 'commissions/application_required_logs.html'
    insertion = get_object_or_404(
        ApplicationInsertionRequired,
        application=application,
        pk=insertion_pk
    )

    logs = Log.objects.filter(
        content_type=ContentType.objects.get_for_model(insertion),
        object_id=insertion.pk
    )

    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            'logs': logs,
            'target_teaching': target_teaching
        }
    )


@login_required
@belongs_to_commission
@application_check
@application_check
def application_free_review_logs(request, call_pk, application_pk, year, insertion_pk, commission=None, application=None):
    free_credits_rule = get_object_or_404(
        CallFreeCreditsRule,
        is_active=True,
        call=application.call,
        course_year=year
    )

    insertion = get_object_or_404(
        ApplicationInsertionFree,
        application=application,
        pk=insertion_pk
    )

    logs = Log.objects.filter(
        content_type=ContentType.objects.get_for_model(insertion),
        object_id=insertion.pk
    )
    template = 'commissions/application_free_logs.html'

    return render(
        request,
        template,
        {
            'application': application,
            'commission': commission,
            'free_credits_rule': free_credits_rule,
            'logs': logs,
        }
    )


@login_required
@belongs_to_commission
@application_check
def export_xls(request, call_pk, application_pk, commission=None, application=None):
    # Crea un nuovo file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Dati"

    ws.append(["Candidato", str(application.user)])
    ws.append(["Nazionalità", application.user_country])
    ws.append(["Università", f"{application.home_university} ({application.home_city} - {application.home_country})"])
    ws.append(["Corso", application.home_course])

    ws.append([])

    ws.append(["Data di invio", application.submission_date.strftime("%d/%m/%Y %H:%M:%S")])
    ws.append(["Num. protocollo", application.protocol_number])
    ws.append(["Data protocollo", application.protocol_date.strftime("%d/%m/%Y %H:%M:%S") if application.protocol_date else "-"])

    # required
    insertions_required = ApplicationInsertionRequired.objects.filter(
        application=application
    ).prefetch_related('review').order_by('target_teaching_name')

    if insertions_required.exists():
        ws.append([])

        ws.append(["Insegnamenti previsti dal piano"])
        # Applica il grassetto
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

        ws.append([])

        labels = [
            "Attività formativa sostenuta"
        ]

        if not application.call.insertions_only_from_same_course:
            labels.extend(
                [
                    "Università",
                    "Corso",
                ]
            )

        labels.extend(
            [
                "CFU",
                "Voto",
                "Attività formativa convalidata",
                "CFU riconosciuti",
                "Voto riconosciuto",
                "Note",
            ]
        )

        ws.append(labels)

        # Applica il grassetto all'ultima riga
        for col in range(1, len(labels) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

        for required in insertions_required:
            data = [
                f"{required.source_teaching_name} ({required.source_teaching_ssd or '-'})",
            ]

            if not application.call.insertions_only_from_same_course:
                data.extend(
                    [
                        f"{required.source_university} - {required.source_university_city} ({required.source_university_city})",
                        required.source_degree_course
                    ]
                )

            data.extend(
                [
                    required.source_teaching_credits,
                    required.source_teaching_grade,
                    f"{required.target_teaching_cod} -{required.target_teaching_name} - {required.target_teaching_ssd} ({required.target_teaching_credits} CFU)",
                    required.review.changed_credits if hasattr(required, 'review') else required.target_teaching_credits,
                    required.review.changed_grade if hasattr(required, 'review') else required.source_teaching_grade,
                    required.review.notes if hasattr(required, 'review') else "-",
                ]
            )

            ws.append(data)
    # end required

    # free
    insertions_free = ApplicationInsertionFree.objects.filter(
        application=application
    ).select_related('free_credits')

    if insertions_free.exists():
        ws.append([])
        ws.append(["Insegnamenti a scelta"])
        # Applica il grassetto
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append([])

        labels = [
            "Attività formativa sostenuta",
            "Università",
            "Corso",
            "CFU",
            "Voto",
            "Vincolo insegnamenti a scelta",
            "CFU riconosciuti",
            "Voto riconosciuto",
            "Note",
        ]

        ws.append(labels)

        # Applica il grassetto all'ultima riga
        for col in range(1, len(labels) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

        for free in insertions_free:
            data = [
                f"{free.source_teaching_name} ({free.source_teaching_ssd or '-'})",
                f"{free.source_university} - {free.source_university_city} ({free.source_university_city})",
                free.source_degree_course,
                free.source_teaching_credits,
                free.source_teaching_grade,
                f"Insegnamenti a scelta {free.free_credits.course_year}° anno (max {free.free_credits.max_value} CFU)",
                free.review.changed_credits if hasattr(free, 'review') else free.target_teaching_credits,
                free.review.changed_grade if hasattr(free, 'review') else free.source_teaching_grade,
                free.review.notes if hasattr(free, 'review') else "-",
            ]
            ws.append(data)
    # end free

    # Prepara la risposta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="domanda_{application.pk}.xlsx"'

    # Salva il file direttamente nella response
    wb.save(response)

    return response
